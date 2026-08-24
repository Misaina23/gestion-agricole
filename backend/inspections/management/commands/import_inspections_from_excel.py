"""
Management command to import inspection data from the T06 Excel file.

Inspection records are derived from:
- Sheet 2 -> last internal inspection date + inspector
- Sheet 3 -> production unit last internal inspection date + inspector

Existing Inspection objects are matched by:
- producer + inspection_type + planned_date/actual_date + inspector
"""
import os
import secrets
from datetime import datetime

from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model

from producers.models import Producer
from parcels.models import Parcel
from inspections.models import Inspection

User = get_user_model()


def to_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_str(value):
    if value is None:
        return ''
    return str(value).strip()


class Command(BaseCommand):
    help = 'Import inspection data from the T06 cooperative Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Do not save to database')
        parser.add_argument('--limit', type=int, default=None, help='Limit number of producer rows processed')

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f'File not found: {excel_path}'))
            return

        dry_run = options['dry_run']
        limit = options.get('limit')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN - no data will be saved'))

        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws_members = wb['2-Registre des membres']
        ws_units = wb['3-Registre des unités'] if '3-Registre des unités' in wb.sheetnames else None

        self.stdout.write('Loading reference data...')
        producers = {p.code: p for p in Producer.objects.all()}
        producers_by_id = {p.id: p for p in Producer.objects.all()}
        parcels_by_producer = {}
        for parcel in Parcel.objects.select_related('producer').all():
            parcels_by_producer.setdefault(parcel.producer_id, []).append(parcel)

        existing_inspections = set(
            Inspection.objects.values_list('code', flat=True)
        )
        inspector_cache = {}

        def get_or_create_inspector(name):
            name = to_str(name)
            if not name:
                return None
            key = name.lower()
            if key in inspector_cache:
                return inspector_cache[key]
            inspector = User.objects.filter(last_name__iexact=name).first()
            if not inspector:
                inspector = User.objects.filter(first_name__iexact=name).first()
            if not inspector:
                inspector = User.objects.filter(username__iexact=name).first()
            if not inspector:
                inspector = User.objects.create_user(
                    username=name.lower().replace(' ', '.'),
                    password=secrets.token_urlsafe(16),
                    last_name=name,
                    first_name='',
                    email=f'{name.lower().replace(" ", ".")}@inspector.local',
                    is_field_agent=False,
                    is_staff=True,
                )
            inspector_cache[key] = inspector
            return inspector

        def build_code(producer_code, inspection_type, date_value):
            return f"INS-{producer_code}-{inspection_type}-{date_value}".upper().replace(' ', '-')[:50]

        to_create = []
        to_update = []
        skipped = 0
        processed = 0

        self.stdout.write('Processing member sheet...')
        rows = list(ws_members.iter_rows(min_row=3, values_only=True))
        for row in rows:
            if limit is not None and processed >= limit:
                break
            producer_code = to_str(row[3])
            producer = producers.get(producer_code)
            if not producer:
                skipped += 1
                processed += 1
                continue

            internal_date = to_date(row[12])
            internal_inspector_name = to_str(row[13])
            external_date = to_date(row[14])

            inspector = get_or_create_inspector(internal_inspector_name) if internal_inspector_name else None
            parcel = (parcels_by_producer.get(producer.id) or [None])[0]

            if internal_date:
                code = build_code(producer.code, 'routine', internal_date)
                if code in existing_inspections:
                    updated += 1
                else:
                    to_create.append(Inspection(
                        code=code,
                        producer=producer,
                        parcel=parcel,
                        inspection_type='routine',
                        planned_date=internal_date,
                        actual_date=internal_date,
                        inspector=inspector,
                        status='completed',
                        result='passed',
                        observations='Inspection interne importée depuis le registre des membres.',
                        notes='',
                    ))
                    existing_inspections.add(code)

            if external_date:
                code = build_code(producer.code, 'certification', external_date)
                if code in existing_inspections:
                    updated += 1
                else:
                    to_create.append(Inspection(
                        code=code,
                        producer=producer,
                        parcel=parcel,
                        inspection_type='certification',
                        planned_date=external_date,
                        actual_date=external_date,
                        inspector=inspector,
                        status='completed',
                        result='passed',
                        observations='Inspection ECOCERT importée depuis le registre des membres.',
                        notes='',
                    ))
                    existing_inspections.add(code)

            processed += 1
            if processed % 1000 == 0:
                self.stdout.write(f'  ... {processed} rows processed')

        self.stdout.write(f'Member sheet done. Pending creates: {len(to_create)}')

        # Sheet 3 -> units
        if ws_units:
            self.stdout.write('Processing unit sheet...')
            processed_units = 0
            for row in ws_units.iter_rows(min_row=3, values_only=True):
                if limit is not None and processed_units >= limit:
                    break
                unit_name = to_str(row[0])
                unit_code = to_str(row[1])
                if not unit_name and not unit_code:
                    continue

                internal_date = to_date(row[12])
                internal_inspector_name = to_str(row[13])
                external_date = to_date(row[14])

                inspector = get_or_create_inspector(internal_inspector_name) if internal_inspector_name else None
                producer = None
                if unit_name:
                    suffix = unit_name.split('/')[-1].strip()
                    for p in producers.values():
                        if suffix and (suffix in (p.unit_name or '')):
                            producer = p
                            break

                if not producer:
                    skipped += 1
                    processed_units += 1
                    continue

                parcel = (parcels_by_producer.get(producer.id) or [None])[0]

                if internal_date:
                    code = build_code(producer.code, 'routine', internal_date)
                    if code not in existing_inspections:
                        to_create.append(Inspection(
                            code=code,
                            producer=producer,
                            parcel=parcel,
                            inspection_type='routine',
                            planned_date=internal_date,
                            actual_date=internal_date,
                            inspector=inspector,
                            status='completed',
                            result='passed',
                            observations='Inspection interne importée depuis le registre des unités.',
                            notes='',
                        ))
                        existing_inspections.add(code)

                if external_date:
                    code = build_code(producer.code, 'certification', external_date)
                    if code not in existing_inspections:
                        to_create.append(Inspection(
                            code=code,
                            producer=producer,
                            parcel=parcel,
                            inspection_type='certification',
                            planned_date=external_date,
                            actual_date=external_date,
                            inspector=inspector,
                            status='completed',
                            result='passed',
                            observations='Inspection ECOCERT importée depuis le registre des unités.',
                            notes='',
                        ))
                        existing_inspections.add(code)

                processed_units += 1

        self.stdout.write(f'Unit sheet done. Pending creates: {len(to_create)}')

        if dry_run:
            self.stdout.write(self.style.WARNING(
                f'DRY RUN complete: would create {len(to_create)} inspections, {updated} would be skipped as existing, {skipped} skipped rows.'
            ))
        else:
            self.stdout.write('Saving inspections...')
            with transaction.atomic():
                if to_create:
                    Inspection.objects.bulk_create(to_create, batch_size=1000)
            limit_msg = f' (limited to {limit} rows)' if limit is not None else ''
            self.stdout.write(self.style.SUCCESS(
                f'Import complete{limit_msg}: {len(to_create)} inspections created, {updated} skipped as existing, {skipped} skipped rows.'
            ))
