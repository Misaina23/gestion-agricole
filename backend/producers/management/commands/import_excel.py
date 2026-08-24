"""
Management command to import real cooperative data from the T06 Excel file.

Usage:
    python manage.py import_excel <path_to_excel_file>

Example:
    python manage.py import_excel "T06COOPERATIVE VINTSY ANNEE 2026.-0.xlsx"
"""
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from core.models import Region, District, Commune, Fokontany, ProductionUnit
from producers.models import Producer
from parcels.models import Parcel


def to_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(',', '.')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def to_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def to_str(value):
    if value is None:
        return ''
    return str(value).strip()


class Command(BaseCommand):
    help = 'Import producers and parcels from the T06 cooperative Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Do not save to database')
        parser.add_argument('--clear', action='store_true', help='Clear existing data before import')

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f'File not found: {excel_path}'))
            return

        dry_run = options['dry_run']
        clear = options['clear']

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN - no data will be saved'))

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        ws_members = wb['2-Registre des membres']
        ws_parcels = wb['4-Registre des parcelles']
        ws_units = wb.worksheets[2] if len(wb.sheetnames) > 2 else None

        if clear and not dry_run:
            self.stdout.write('Clearing existing parcels, producers and production units...')
            Parcel.objects.all().delete()
            Producer.objects.all().delete()
            if ws_units:
                ProductionUnit.objects.all().delete()

        # Cache geographical objects
        region_cache = {}
        district_cache = {}
        commune_cache = {}
        producer_cache = {}
        unit_cache = {}

        def get_or_create_region(name):
            if not name:
                return None
            name = name.strip()
            if name in region_cache:
                return region_cache[name]
            code = name[:10].upper().replace(' ', '_')
            try:
                region, _ = Region.objects.get_or_create(name=name, defaults={'code': code, 'is_active': True})
            except Exception:
                region = Region.objects.filter(code=code).first() or Region.objects.create(name=name, code=code, is_active=True)
            region_cache[name] = region
            return region

        def get_or_create_district(name, region):
            if not name or not region:
                return None
            key = (name.strip(), region.id)
            if key in district_cache:
                return district_cache[key]
            code = name[:20].upper().replace(' ', '_')
            try:
                district, _ = District.objects.get_or_create(name=name.strip(), region=region, defaults={'code': code})
            except Exception:
                district = District.objects.filter(code=code, region=region).first() or District.objects.create(name=name.strip(), region=region, code=code)
            district_cache[key] = district
            return district

        def get_or_create_commune(name, region, district=None):
            if not name:
                return None
            key = (name.strip(), region.id, district.id if district else None)
            if key in commune_cache:
                return commune_cache[key]
            code = name[:20].upper().replace(' ', '_')
            try:
                commune, _ = Commune.objects.get_or_create(name=name.strip(), region=region, district=district, defaults={'code': code})
            except Exception:
                commune = Commune.objects.filter(code=code, region=region, district=district).first() or Commune.objects.create(name=name.strip(), region=region, district=district, code=code)
            commune_cache[key] = commune
            return commune

        def get_or_create_producer(code, defaults):
            if code in producer_cache:
                return producer_cache[code]
            producer, _ = Producer.objects.get_or_create(code=code, defaults=defaults)
            producer_cache[code] = producer
            return producer

        def get_or_create_unit(code, defaults):
            if code in unit_cache:
                return unit_cache[code]
            unit, _ = ProductionUnit.objects.get_or_create(code=code, defaults=defaults)
            unit_cache[code] = unit
            return unit

        # --- Import Producers (Sheet 2) ---
        self.stdout.write('Importing producers...')
        member_rows = 0
        producers_created = 0
        producers_updated = 0

        for row in ws_members.iter_rows(min_row=3, values_only=True):
            if not to_str(row[3]):
                continue
            member_rows += 1

            code = to_str(row[3])
            unit_name = to_str(row[0]) or None
            last_name = to_str(row[1]) or code
            first_name = to_str(row[2]) or None
            phone = to_str(row[4]) or None
            joined_at = to_date(row[5])
            risk_category = to_str(row[8]).lower() if row[8] else None
            if risk_category not in ('low', 'medium', 'high'):
                risk_category = None
            identified_risks = to_str(row[9]) or None
            member_processing = to_str(row[11]).lower() if row[11] else None
            if member_processing not in ('yes', 'no'):
                member_processing = None
            processing_activities = to_str(row[12]) or None
            last_internal_inspection_at = to_date(row[13])
            internal_inspector_name = to_str(row[14]) or None
            last_external_inspection_at = to_date(row[15])
            eu_status = to_str(row[16]).lower() if row[16] else None
            if eu_status not in ('active', 'suspended', 'withdrawn', 'abandoned'):
                eu_status = 'active' if to_str(row[16]) else None
            nop_status = to_str(row[17]).lower() if row[17] else None
            if nop_status not in ('active', 'suspended', 'abandoned'):
                nop_status = 'active' if to_str(row[17]) else None
            exclusion_reason = to_str(row[18]) or None
            exclusion_date = to_date(row[19])

            region = get_or_create_region(unit_name.split('/')[0].strip() if unit_name else None)
            district = get_or_create_district(unit_name.split('/')[1].strip() if unit_name and '/' in unit_name else None, region)
            commune = get_or_create_commune(unit_name.split('/')[-1].strip() if unit_name else None, region, district)

            if dry_run:
                producers_created += 1
                continue

            producer, created = Producer.objects.update_or_create(
                code=code,
                defaults={
                    'last_name': last_name,
                    'first_name': first_name,
                    'unit_name': unit_name,
                    'region': region,
                    'district': district,
                    'commune': commune,
                    'phone': phone,
                    'joined_at': joined_at,
                    'risk_category': risk_category,
                    'identified_risks': identified_risks,
                    'member_processing': member_processing,
                    'processing_activities': processing_activities,
                    'last_internal_inspection_at': last_internal_inspection_at,
                    'internal_inspector_name': internal_inspector_name,
                    'last_external_inspection_at': last_external_inspection_at,
                    'eu_status': eu_status,
                    'nop_status': nop_status,
                    'exclusion_reason': exclusion_reason,
                    'exclusion_date': exclusion_date,
                    'status': 'active' if eu_status == 'active' and nop_status == 'active' else 'inactive',
                    'synced': True,
                },
            )
            if created:
                producers_created += 1
            else:
                producers_updated += 1

            if member_rows % 200 == 0:
                self.stdout.write(f'  ... {member_rows} producers processed')

        self.stdout.write(f'  Processed {member_rows} producer rows ({producers_created} created, {producers_updated} updated)')

        # --- Import Production Units (Sheet 3) ---
        unit_rows = 0
        units_created = 0
        units_updated = 0
        if ws_units:
            self.stdout.write('Importing production units...')
            for row in ws_units.iter_rows(min_row=3, values_only=True):
                if not to_str(row[0]) and not to_str(row[1]):
                    continue
                unit_rows += 1

                unit_name = to_str(row[0]) or None
                unit_code = to_str(row[1]) or None
                if not unit_code:
                    continue

                unit_type = 'group'
                if unit_name:
                    lowered = unit_name.lower()
                    if 'site' in lowered:
                        unit_type = 'site'
                    elif 'village' in lowered:
                        unit_type = 'village'
                    elif 'region' in lowered:
                        unit_type = 'region'

                region = get_or_create_region(unit_name.split('/')[0].strip() if unit_name else None)
                district = get_or_create_district(unit_name.split('/')[1].strip() if unit_name and '/' in unit_name else None, region)
                commune = get_or_create_commune(unit_name.split('/')[-1].strip() if unit_name else None, region, district)

                defaults = {
                    'name': unit_name or unit_code,
                    'unit_type': unit_type,
                    'region': region,
                    'district': district,
                    'commune': commune,
                    'manager_name': to_str(row[2]) or None,
                    'manager_function': to_str(row[3]) or None,
                    'phone': to_str(row[4]) or None,
                    'email': to_str(row[5]) or None,
                    'members_count': to_int(row[6]) or 0,
                    'total_area': to_decimal(row[7]) or Decimal('0'),
                    'creation_date': to_date(row[8]),
                    'status': 'active',
                    'notes': to_str(row[10]) or None,
                }

                if dry_run:
                    units_created += 1
                    continue

                unit, created = ProductionUnit.objects.update_or_create(
                    code=unit_code,
                    defaults=defaults,
                )
                if created:
                    units_created += 1
                else:
                    units_updated += 1

            self.stdout.write(f'  Processed {unit_rows} unit rows ({units_created} created, {units_updated} updated)')
        else:
            self.stdout.write('Sheet 3 not found, skipping production units import.')

        # --- Import Parcels (Sheet 4) ---
        self.stdout.write('Importing parcels...')
        parcel_rows = 0
        parcels_created = 0
        parcels_updated = 0
        errors = []

        # Preload existing parcels for faster lookup
        if not dry_run:
            existing_parcels = {
                (p.producer_id, p.code): p for p in Parcel.objects.all()
            }
        else:
            existing_parcels = {}

        new_parcels = []
        updated_parcels = []

        for row in ws_parcels.iter_rows(min_row=3, values_only=True):
            if not to_str(row[0]):
                continue
            parcel_rows += 1

            producer_code = to_str(row[0])
            parcel_code = to_str(row[3]) or 'P1'
            registration_date = to_date(row[4])
            area = to_decimal(row[5])
            main_crop = to_str(row[6]) or None
            intercrop = to_str(row[7]) or None
            vanilla_plants = to_int(row[8]) or 0
            bio_location = to_str(row[9]).lower() if row[9] else None
            if bio_location not in ('oui', 'non', 'yes', 'no'):
                bio_location = None
            latitude = to_decimal(row[10])
            longitude = to_decimal(row[11])
            conversion_start_date = to_date(row[12])
            conversion_status = to_str(row[13]).lower() if row[13] else None
            if conversion_status not in ('biologique', 'en conversion', 'conventionnelle', 'organic', 'conversion', 'conventional'):
                conversion_status = None
            if conversion_status == 'biologique':
                conversion_status = 'organic'
            elif conversion_status == 'en conversion':
                conversion_status = 'conversion'
            elif conversion_status == 'conventionnelle':
                conversion_status = 'conventional'
            last_used_date = to_date(row[14])
            eu_status = to_str(row[15])
            nop_status = to_str(row[16])
            estimated_yield = to_decimal(row[17])
            actual_harvest = to_decimal(row[18])
            delivered_quantity = to_decimal(row[20])

            try:
                producer = Producer.objects.get(code=producer_code)
            except Producer.DoesNotExist:
                errors.append(f'Producer not found: {producer_code} (parcel {parcel_code})')
                continue

            key = (producer.id, parcel_code)
            defaults = {
                'registration_date': registration_date,
                'area': area or Decimal('0'),
                'main_crop': main_crop,
                'intercrop': intercrop,
                'vanilla_plants': vanilla_plants,
                'bio_location': bio_location,
                'latitude': latitude,
                'longitude': longitude,
                'conversion_start_date': conversion_start_date,
                'conversion_status': conversion_status or 'organic',
                'last_used_date': last_used_date,
                'eu_status': eu_status,
                'nop_status': nop_status,
                'estimated_yield': estimated_yield,
                'actual_harvest': actual_harvest,
                'delivered_quantity': delivered_quantity,
                'synced': True,
            }

            if dry_run:
                parcels_created += 1
                continue

            existing = existing_parcels.get(key)
            if existing:
                for attr, value in defaults.items():
                    setattr(existing, attr, value)
                updated_parcels.append(existing)
                parcels_updated += 1
            else:
                new_parcel = Parcel(producer=producer, code=parcel_code, **defaults)
                new_parcels.append(new_parcel)
                parcels_created += 1

            if parcel_rows % 2000 == 0:
                self.stdout.write(f'  ... {parcel_rows} parcels processed')

        if not dry_run:
            if new_parcels:
                Parcel.objects.bulk_create(new_parcels, batch_size=1000)
            if updated_parcels:
                Parcel.objects.bulk_update(updated_parcels, list(defaults.keys()), batch_size=1000)

        self.stdout.write(f'  Processed {parcel_rows} parcel rows ({parcels_created} created, {parcels_updated} updated)')

        if errors:
            self.stdout.write(self.style.WARNING(f'\n{len(errors)} errors:'))
            for err in errors[:20]:
                self.stdout.write(f'  - {err}')
            if len(errors) > 20:
                self.stdout.write(f'  ... and {len(errors) - 20} more')

        self.stdout.write(self.style.SUCCESS(
            f'\nImport complete: {producers_created} producers created, {producers_updated} updated, '
            f'{parcels_created} parcels created, {parcels_updated} updated.'
        ))

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN - no changes were saved.'))
