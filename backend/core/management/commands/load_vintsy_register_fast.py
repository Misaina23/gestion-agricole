"""Bulk-load the approved Vintsy T06 workbook over a high-latency database link."""
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Region, District, Commune, Fokontany
from producers.models import Producer
from parcels.models import Parcel, ParcelRegisterHarvest
from .import_vintsy_register import value, date_value, decimal_value, conversion_status


class Command(BaseCommand):
    help = 'Bulk-load real Vintsy members and parcels from T06 Excel.'

    def add_arguments(self, parser):
        parser.add_argument('workbook', type=Path)

    @transaction.atomic
    def handle(self, *args, **options):
        path = options['workbook']
        if not path.is_file():
            raise CommandError(f'Classeur introuvable : {path}')
        wb = load_workbook(path, read_only=True, data_only=True)
        members, parcel_sheet = wb['2-Registre des membres'], wb['4-Registre des parcelles']
        region, _ = Region.objects.get_or_create(code='VINTSY', defaults={'name': 'Coopérative Vintsy'})
        district, _ = District.objects.get_or_create(code='VINTSY', defaults={'name': 'Coopérative Vintsy', 'region': region})
        commune, _ = Commune.objects.get_or_create(code='VINTSY-NR', defaults={'name': 'Non renseigné', 'region': region, 'district': district})
        fokontanys, producer_rows = {}, []
        for row in members.iter_rows(min_row=3, values_only=True):
            code = value(row[3])
            if not code:
                continue
            zone = value(row[0])
            if zone and zone not in fokontanys:
                fcode = 'VINTSY-' + re.sub(r'[^A-Z0-9]+', '-', zone.upper()).strip('-')[:20]
                fokontanys[zone], _ = Fokontany.objects.get_or_create(code=fcode, defaults={'name': zone[:100], 'commune': commune})
            producer_rows.append((code, row, zone))
        known_codes = set(Producer.objects.filter(code__in=[item[0] for item in producer_rows]).values_list('code', flat=True))
        Producer.objects.bulk_create([
            Producer(code=code, name=' '.join(part for part in (value(row[1]), value(row[2])) if part) or code,
                phone=value(row[4]), joined_at=date_value(row[5]), region=region, district=district, commune=commune,
                fokontany=fokontanys.get(zone), address=zone,
                status='active' if (value(row[16]) or '').lower() == 'actif' else 'inactive', risk_category=value(row[8]),
                identified_risks=value(row[9]), member_processing=value(row[11]), processing_activities=value(row[12]),
                last_internal_inspection_at=date_value(row[13]), internal_inspector_name=value(row[14]),
                last_external_inspection_at=date_value(row[15]), eu_status=value(row[16]), nop_status=value(row[17]))
            for code, row, zone in producer_rows if code not in known_codes
        ], batch_size=500)
        producers = Producer.objects.in_bulk([item[0] for item in producer_rows], field_name='code')
        existing = set(Parcel.objects.filter(producer__code__in=producers.keys()).values_list('producer__code', 'code'))
        new_parcels = []
        for row in parcel_sheet.iter_rows(min_row=3, values_only=True):
            pcode, code, area = value(row[0]), value(row[3]), decimal_value(row[5])
            if pcode not in producers or not code or area is None or (pcode, code) in existing:
                continue
            state, raw = conversion_status(row[13]), value(row[13]) or ''
            match = re.search(r'\b(C[123])\b', raw.upper())
            new_parcels.append(Parcel(producer=producers[pcode], code=code, name=code, area=area, main_crop=value(row[6]), intercrop=value(row[7]),
                vanilla_plants=int(decimal_value(row[8]) or 0), latitude=decimal_value(row[10], True), longitude=decimal_value(row[11], True),
                conversion_start_date=date_value(row[12]), conversion_status=state, conversion_level=match.group(1) if state == 'conversion' and match else None,
                eu_status=value(row[13]), nop_status=value(row[15]), estimated_yield=decimal_value(row[16]), actual_harvest=decimal_value(row[17]),
                delivered_quantity=decimal_value(row[19]), status='active', is_certified=state == 'organic'))
        Parcel.objects.bulk_create(new_parcels, batch_size=500, ignore_conflicts=True)
        parcel_lookup = {
            (p.producer.code, p.code): p
            for p in Parcel.objects.filter(producer__code__in=producers.keys()).select_related('producer')
        }
        register_harvests = []
        # T06 columns: current main 17-20, intercrop 2 22-25, intercrop 1
        # 27-30; N-1 stores harvest/delivery only in columns 31-36.
        slots = [('main', 16), ('intercrop_2', 21), ('intercrop_1', 26)]
        for row in parcel_sheet.iter_rows(min_row=3, values_only=True):
            parcel = parcel_lookup.get((value(row[0]), value(row[3])))
            if not parcel:
                continue
            for slot, offset in slots:
                fields = [decimal_value(row[offset + index]) for index in range(4)]
                if any(item is not None for item in fields):
                    register_harvests.append(ParcelRegisterHarvest(parcel=parcel, period='current', crop_slot=slot,
                        estimated_yield=fields[0], actual_harvest=fields[1], actual_yield=fields[2], delivered_quantity=fields[3]))
            for slot, offset in [('main', 30), ('intercrop_1', 32), ('intercrop_2', 34)]:
                harvested, delivered = decimal_value(row[offset]), decimal_value(row[offset + 1])
                if harvested is not None or delivered is not None:
                    register_harvests.append(ParcelRegisterHarvest(parcel=parcel, period='previous', crop_slot=slot,
                        actual_harvest=harvested, delivered_quantity=delivered))
        ParcelRegisterHarvest.objects.bulk_create(register_harvests, batch_size=500, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(f'Chargement terminé : {len(producer_rows)} producteurs source, {len(new_parcels)} parcelles et {len(register_harvests)} relevés récolte ajoutés.'))
