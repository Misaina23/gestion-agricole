"""Import the cooperative's T06 register without manufacturing data.

The workbook uses a member code as the producer key and a parcel number that
is only unique inside that producer.  Running this command again updates the
same records; it never creates a second producer for another parcel row.
"""
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Region, District, Commune, Fokontany
from producers.models import Producer
from parcels.models import Parcel


def value(cell):
    if cell is None:
        return None
    text = str(cell).strip()
    return text or None


def date_value(cell):
    if not cell:
        return None
    if hasattr(cell, 'date'):
        return cell.date()
    for pattern in ('%d/%m/%Y', '%d.%m.%Y', '%d/%m/%y', '%d.%m.%y'):
        try:
            return datetime.strptime(str(cell).strip(), pattern).date()
        except ValueError:
            continue
    return None


def decimal_value(cell, coordinate=False):
    if cell in (None, ''):
        return None
    text = str(cell).strip().replace(' ', '').replace(',', '.')
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    # Some longitude cells have lost their decimal separator in Excel (e.g.
    # 4952325 means 49.52325); preserve valid values and repair only this
    # unambiguous presentation defect.
    if coordinate and abs(number) > 180 and re.fullmatch(r'-?\d+', text):
        number /= Decimal('100000')
    return number


def conversion_status(raw):
    normalized = (value(raw) or '').lower()
    if 'conversion' in normalized:
        return 'conversion'
    if 'biolog' in normalized:
        return 'organic'
    if 'convention' in normalized:
        return 'conventional'
    return None


class Command(BaseCommand):
    help = 'Import members and parcels from the real Vintsy T06 Excel register.'

    def add_arguments(self, parser):
        parser.add_argument('workbook', type=Path, help='Path to the T06 .xlsx workbook')

    @transaction.atomic
    def handle(self, *args, **options):
        path = options['workbook']
        if not path.is_file():
            raise CommandError(f'Classeur introuvable : {path}')
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            members = workbook['2-Registre des membres']
            parcels = workbook['4-Registre des parcelles']
        except KeyError as exc:
            raise CommandError(f'Feuille obligatoire absente : {exc}') from exc

        region, _ = Region.objects.get_or_create(code='VINTSY', defaults={'name': 'Coopérative Vintsy'})
        district, _ = District.objects.get_or_create(code='VINTSY', defaults={'name': 'Coopérative Vintsy', 'region': region})
        commune, _ = Commune.objects.get_or_create(
            code='VINTSY-NR', defaults={'name': 'Non renseigné', 'region': region, 'district': district}
        )

        imported_producers = 0
        imported_parcels = 0
        producer_by_code = {}
        for row in members.iter_rows(min_row=3, values_only=True):
            code = value(row[3])
            if not code:
                continue
            zone = value(row[0])
            fokontany = None
            if zone:
                safe_code = 'VINTSY-' + re.sub(r'[^A-Z0-9]+', '-', zone.upper()).strip('-')[:20]
                fokontany, _ = Fokontany.objects.get_or_create(code=safe_code, defaults={'name': zone[:100], 'commune': commune})
            name = ' '.join(part for part in (value(row[1]), value(row[2])) if part) or code
            producer, _ = Producer.objects.update_or_create(
                code=code,
                defaults={
                    'name': name, 'phone': value(row[4]), 'joined_at': date_value(row[5]),
                    'region': region, 'district': district, 'commune': commune, 'fokontany': fokontany,
                    'address': zone, 'status': 'active' if (value(row[16]) or '').lower() == 'actif' else 'inactive',
                    'risk_category': value(row[8]), 'identified_risks': value(row[9]),
                    'member_processing': value(row[11]), 'processing_activities': value(row[12]),
                    'last_internal_inspection_at': date_value(row[13]), 'internal_inspector_name': value(row[14]),
                    'last_external_inspection_at': date_value(row[15]), 'eu_status': value(row[16]), 'nop_status': value(row[17]),
                },
            )
            producer_by_code[code] = producer
            imported_producers += 1

        for row in parcels.iter_rows(min_row=3, values_only=True):
            producer = producer_by_code.get(value(row[0]))
            parcel_code = value(row[3])
            area = decimal_value(row[5])
            if not producer or not parcel_code or area is None:
                continue
            status = conversion_status(row[13])
            level = None
            raw_status = value(row[13]) or ''
            level_match = re.search(r'\b(C[123])\b', raw_status.upper())
            if level_match:
                level = level_match.group(1)
            if status != 'conversion':
                level = None
            Parcel.objects.update_or_create(
                producer=producer, code=parcel_code,
                defaults={
                    'name': parcel_code, 'area': area, 'main_crop': value(row[6]), 'intercrop': value(row[7]),
                    'vanilla_plants': int(decimal_value(row[8]) or 0),
                    'latitude': decimal_value(row[10], coordinate=True), 'longitude': decimal_value(row[11], coordinate=True),
                    'conversion_start_date': date_value(row[12]), 'conversion_status': status, 'conversion_level': level,
                    'eu_status': value(row[13]), 'nop_status': value(row[15]),
                    'estimated_yield': decimal_value(row[16]), 'actual_harvest': decimal_value(row[17]),
                    'delivered_quantity': decimal_value(row[19]),
                    'status': 'active', 'is_certified': status == 'organic',
                },
            )
            imported_parcels += 1

        self.stdout.write(self.style.SUCCESS(
            f'Import terminé : {imported_producers} producteurs et {imported_parcels} parcelles synchronisés.'
        ))
