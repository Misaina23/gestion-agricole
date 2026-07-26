import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional


def _get_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Workbook, Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None


def _get_reportlab():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        return (
            colors, A4, SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, getSampleStyleSheet, ParagraphStyle, cm
        )
    except ImportError:
        return None


def generate_excel(headers: List[str], rows: List[List[Any]], sheet_name: str = 'Sheet1') -> bytes:
    """Generate an Excel file from headers and rows. Falls back to CSV if openpyxl not installed."""
    openpyxl_result = _get_openpyxl()
    if openpyxl_result is None:
        return generate_csv(headers, rows).encode('utf-8-sig')
    
    Workbook, Font, PatternFill, Alignment, Border, Side = openpyxl_result
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_csv(headers: List[str], rows: List[List[Any]]) -> str:
    """Generate a CSV string from headers and rows."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def generate_pdf(
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    subtitle: Optional[str] = None,
    generated_by: Optional[str] = None,
) -> bytes:
    """Generate a PDF file with a table. Falls back to CSV if reportlab not installed."""
    openpyxl_result = _get_openpyxl()
    if openpyxl_result is None:
        return generate_csv(headers, rows).encode('utf-8-sig')
    
    reportlab_result = _get_reportlab()
    if reportlab_result is None:
        return generate_csv(headers, rows).encode('utf-8-sig')
    
    (colors, A4, SimpleDocTemplate, Table, TableStyle,
     Paragraph, Spacer, getSampleStyleSheet, ParagraphStyle, cm) = reportlab_result
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=20,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=10,
        alignment=1,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))
    if generated_by:
        elements.append(Paragraph(f"Genere par: {generated_by}", subtitle_style))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.5 * cm))

    table_data = [headers] + rows
    col_widths = [doc.width / len(headers)] * len(headers)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def build_export_response(
    data: List[Dict[str, Any]],
    filename_prefix: str,
    export_format: str = 'xlsx',
) -> tuple:
    """
    Generic export helper: takes a list of dicts and returns (content_type, content, filename).
    """
    if not data:
        return ('application/json', b'[]', f'{filename_prefix}_empty.json')

    headers = list(data[0].keys())
    rows = [[item.get(key, '') for key in headers] for item in data]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_filename = f'{filename_prefix}_{timestamp}'

    if export_format == 'xlsx':
        content = generate_excel(headers, rows, sheet_name=filename_prefix)
        return ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', content, f'{base_filename}.xlsx')
    elif export_format == 'csv':
        content = generate_csv(headers, rows)
        return ('text/csv', content.encode('utf-8-sig'), f'{base_filename}.csv')
    elif export_format == 'pdf':
        content = generate_pdf(
            title=filename_prefix.capitalize(),
            headers=headers,
            rows=rows,
            generated_by='VIDEEKO Platform',
        )
        return ('application/pdf', content, f'{base_filename}.pdf')
    else:
        raise ValueError(f"Format d'export non supporte: {export_format}")
