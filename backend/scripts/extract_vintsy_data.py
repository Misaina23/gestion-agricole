"""Extract the real Vintsy T06 cooperative register into JSON seed data.

This is the *data* half of the seeding pipeline.  It reads the cooperative's
official Excel register (sheets 1-5) and writes faithful JSON files under
``database/data``.  The matching management command ``seed`` consumes those
files and applies the *logic* (relations, mapping, transactions).

Run::

    python backend/scripts/extract_vintsy_data.py

The Excel workbook is never imported by the running application: it is only a
source used here to build the initial dataset.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_WORKBOOK = REPO_ROOT / "T06COOPERATIVE VINTSY ANNEE 2026.-0.xlsx"
DATA_DIR = Path(__file__).resolve().parent.parent / "database" / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)


def value(cell):
    if cell is None:
        return None
    text = str(cell).strip()
    return text or None


def parse_date(cell):
    if not cell:
        return None
    if hasattr(cell, "date"):
        return cell.date().isoformat()
    for pattern in ("%d/%m/%Y", "%d.%m.%Y", "%d/%m/%y", "%d.%m.%y"):
        try:
            return datetime.strptime(str(cell).strip(), pattern).date().isoformat()
        except ValueError:
            continue
    return None


def parse_decimal(cell, coordinate=False):
    if cell in (None, ""):
        return None
    raw = str(cell).strip()
    text = raw.replace(" ", "").replace(",", ".")
    if text.count(".") > 1:
        parts = [part for part in text.split(".") if part]
        text = parts[0] + "." + "".join(parts[1:])
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    # Some GPS cells lost their decimal separator in Excel (e.g. 4953297 should
    # be 49.53297).  Repair only this unambiguous presentation defect.
    if coordinate and abs(number) > 180 and re.fullmatch(r"-?\d+", text):
        number /= Decimal("100000")
    return float(number)


def as_int(cell):
    number = parse_decimal(cell)
    if number is None:
        return None
    return int(number)


def conversion_level(raw_status):
    if not raw_status:
        return None
    match = re.search(r"\b(C[123])\b", str(raw_status).upper())
    return match.group(1) if match else None


def write_json(name, rows):
    path = DATA_DIR / name
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def main(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    members = workbook["2-Registre des membres"]
    parcels = workbook["4-Registre des parcelles"]

    producers = []
    seen_producer = set()
    for row in members.iter_rows(min_row=3, values_only=True):
        code = value(row[3])
        if not code:
            continue
        if code in seen_producer:
            continue
        seen_producer.add(code)
        producers.append(
            {
                "code": code,
                "unit_name": value(row[0]),
                "last_name": value(row[1]),
                "first_name": value(row[2]),
                "phone": value(row[4]),
                "joined_at": parse_date(row[5]),
                "risk_category": value(row[8]),
                "identified_risks": value(row[9]),
                "member_processing": value(row[11]),
                "processing_activities": value(row[12]),
                "last_internal_inspection_at": parse_date(row[13]),
                "internal_inspector_name": value(row[14]),
                "last_external_inspection_at": parse_date(row[15]),
                "eu_status": value(row[16]),
                "nop_status": value(row[17]),
                "exclusion_reason": value(row[18]),
                "exclusion_date": parse_date(row[19]),
            }
        )

    parcel_rows = []
    harvest_rows = []
    seen_parcel = set()
    for row in parcels.iter_rows(min_row=3, values_only=True):
        producer_code = value(row[0])
        parcel_code = value(row[3])
        area = parse_decimal(row[5])
        if not producer_code or not parcel_code or area is None:
            continue
        key = (producer_code, parcel_code)
        if key in seen_parcel:
            continue
        seen_parcel.add(key)
        parcel_rows.append(
            {
                "producer_code": producer_code,
                "code": parcel_code,
                "registration_date": parse_date(row[4]),
                "area": area,
                "main_crop": value(row[6]),
                "intercrop": value(row[7]),
                "vanilla_plants": as_int(row[8]),
                "bio_location": value(row[9]),
                "latitude": parse_decimal(row[10], coordinate=True),
                "longitude": parse_decimal(row[11], coordinate=True),
                "conversion_start_date": parse_date(row[12]),
                "conversion_start_date_raw": value(row[12]),
                "conversion_status": value(row[13]),
                "conversion_level": conversion_level(row[13]),
                "last_used_date": parse_date(row[14]),
                "eu_status": value(row[13]),
                "nop_status": value(row[15]),
                "estimated_yield": parse_decimal(row[16]),
                "actual_harvest": parse_decimal(row[17]),
                "delivered_quantity": parse_decimal(row[19]),
            }
        )

        # Harvest / delivery values exactly as recorded in the register.
        def add_harvest(period, slot, est_col, harvest_col, yield_col, delivered_col):
            est = parse_decimal(row[est_col]) if est_col is not None else None
            harvest = parse_decimal(row[harvest_col])
            yld = parse_decimal(row[yield_col]) if yield_col is not None else None
            delivered = parse_decimal(row[delivered_col])
            if est is None and harvest is None and yld is None and delivered is None:
                return
            harvest_rows.append(
                {
                    "producer_code": producer_code,
                    "parcel_code": parcel_code,
                    "period": period,
                    "crop_slot": slot,
                    "estimated_yield": est,
                    "actual_harvest": harvest,
                    "actual_yield": yld,
                    "delivered_quantity": delivered,
                }
            )

        add_harvest("current", "main", 16, 17, 18, 19)
        add_harvest("current", "intercrop_2", 21, 22, 23, 24)
        add_harvest("current", "intercrop_1", 26, 27, 28, 29)
        add_harvest("previous", "main", 30, 31, None, 31)
        add_harvest("previous", "intercrop_1", 32, 33, None, 33)
        add_harvest("previous", "intercrop_2", 34, 35, None, 35)

    p_producers = write_json("producers.json", producers)
    p_parcels = write_json("parcels.json", parcel_rows)
    p_harvest = write_json("parcel_harvests.json", harvest_rows)

    print(f"Wrote {len(producers)} producers -> {p_producers}")
    print(f"Wrote {len(parcel_rows)} parcels   -> {p_parcels}")
    print(f"Wrote {len(harvest_rows)} harvests -> {p_harvest}")


if __name__ == "__main__":
    import sys

    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not path.is_file():
        raise SystemExit(f"Classeur introuvable : {path}")
    main(path)
